import os
import random
import openpyxl

def getFilesWithExtension(path,ext):
  fileExt = [] # Daftar file yang sesuai dengan extensi
  ext = list(ext) # Extensi File
  files = os.listdir(path) # List File

  for file in files:
   if os.path.splitext(file)[1] in ext:
     fileExt.append(os.path.join(path,file))

  return fileExt

def findFristFileExtension(path,ext):
  for dirpath, dirnames, filenames in os.walk(path):
    for filename in filenames:
      if os.path.splitext(filename)[1] in ext:
        return os.path.join(dirpath, filename)

def randomDataXLSX(nama_file):
  workbook = openpyxl.load_workbook(nama_file)
  data = []

  while True:
    sheet = workbook[random.choice(workbook.sheetnames)]

    for i in sheet[random.randint(1,sheet.max_row)][2:]:
      i = i.value
      if i is not None: data.append(i)
    else:
      if len(data) > 0: break

  return ' '.join(data)
